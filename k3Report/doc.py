# -*- coding: utf-8 -*-
__author__ = 'Виноградов А.Г. г.Белгород  август 2015'

import openpyxl
import os

from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, NamedStyle
from openpyxl.utils.units import cm_to_EMU, EMU_to_inch, pixels_to_points


class DocOpenpyxl:
    """Создание документа Excel
        ExcelDoc - создаёт объект Excel
        newheet - создаёт страницу в книге с заданным именем
        save - сохраняет документ в указанное место
        quit - закрывает документ и объект Excel
        writeval - добавляет запись в документ
    """
    def __init__(self):
        """Создание объекта Excel, инициализация умолчаний"""
        self.xl = openpyxl
        self.wb = self.xl.Workbook()
        self.ws = self.wb.active
        self.PORTRAIT = self.ws.ORIENTATION_PORTRAIT
        self.LANDSCAPE = self.ws.ORIENTATION_LANDSCAPE
        self.first = True
        
        self.sheetorient = self.PORTRAIT
        self.papersize = self.ws.PAPERSIZE_A4
        self.rightmargin = 1.8
        self.leftmargin = 1.8
        self.bottommargin = 1.9
        self.topmargin = 1.9
        
        self.centerhorizontally = 0
        self.font = "Arial Narrow"
        self.fontsize = 12
        self.displayzeros = True
        
        self.generatestyles()
   
    def newsheet(self, name, tabcolor=1):
        """Добавление листа"""
        
        if self.first:
            self.ws.title = name[:31]
            self.first = False
        else:
            self.ws = self.wb.create_sheet(name[:31])

        self.ws.page_setup.orientation = self.sheetorient
        self.ws.page_setup.paperSize = self.papersize
        self.ws.page_margins.left = EMU_to_inch(cm_to_EMU(self.leftmargin))
        self.ws.page_margins.right = EMU_to_inch(cm_to_EMU(self.rightmargin))
        self.ws.page_margins.top = EMU_to_inch(cm_to_EMU(self.topmargin))
        self.ws.page_margins.bottom = EMU_to_inch(cm_to_EMU(self.bottommargin))
        self.ws.print_options.horizontalCentered = self.centerhorizontally

    def generatestyles(self):
        """Создание базовых стилей"""
        styles=[{'name': 'Заголовок 1', 'sc': 'D9D9D9', 'ec': 'D9D9D9', 'bc': '595959', 'bold': True, 'horAlign':'right', 'bb': 'thin', 'ft': 'solid'},
                {'name': 'Заголовок 2', 'sc': '8DB4E2', 'ec': '8DB4E2', 'bc': '16365C', 'bold': True, 'horAlign':'right', 'bb': 'thin', 'ft': 'solid'},
                {'name': 'Заголовок 3', 'sc': 'B8CCE4', 'ec': 'B8CCE4', 'bc': '366092', 'bold': True, 'horAlign':'right', 'bb': 'thin', 'ft': 'solid'},
                {'name': 'Заголовок 4', 'sc': 'E6B8B7', 'ec': 'E6B8B7', 'bc': '963634', 'bold': True, 'horAlign':'right', 'bb': 'thin', 'ft': 'solid'},
                {'name': 'Заголовок 5', 'sc': 'D8E4BC', 'ec': 'D8E4BC', 'bc': '76933C', 'bold': True, 'horAlign':'right', 'bb': 'thin', 'ft': 'solid'},
                {'name': 'Заголовок 6', 'sc': 'CCC0DA', 'ec': 'CCC0DA', 'bc': '60497A', 'bold': True, 'horAlign':'right', 'bb': 'thin', 'ft': 'solid'},
                {'name': 'Заголовок 7', 'sc': 'B7DEE8', 'ec': 'B7DEE8', 'bc': '31869B', 'bold': True, 'horAlign':'right', 'bb': 'thin', 'ft': 'solid'},
                {'name': 'Таблица 1', 'bl': 'thin', 'br': 'thin', 'bt': 'thin', 'bb': 'thin', 'wr': True},
                {'name': 'Шапка 1', 'bl': 'thin', 'br': 'thin', 'bt': 'thin', 'bb': 'thin', 'bold': True}
                ]
        
        for s in styles:
            font = Font(name=self.font, size=self.fontsize, bold=s.get('bold', False), \
                        italic=s.get('italic', False), \
                        vertAlign=s.get('vertAlign'), underline='none', \
                        strike=False, color=s.get('txtclr', '000000') \
                        )
            fill = PatternFill(fill_type=s.get('ft'), start_color=s.get('sc'), end_color=s.get('ec'))

            border = Border(left=Side(border_style=s.get('bl'), color=s.get('bc', '000000')),
                            right=Side(border_style=s.get('br'), color=s.get('bc', '000000')),
                            top=Side(border_style=s.get('bt'), color=s.get('bc', '000000')),
                            bottom=Side(border_style=s.get('bb'), color=s.get('bc', '000000')),
                            diagonal=Side(border_style=None, color=s.get('bc', '000000')),
                            diagonal_direction=0, outline=Side(border_style=None, color=s.get('bc', '000000')),
                            vertical=Side(border_style=s.get('bv'), color=s.get('bc', '000000')),
                            horizontal=Side(border_style=s.get('bh'), color=s.get('bc', '000000'))
                            )
            alignment=Alignment(horizontal=s.get('horAlign', 'left'), \
                                vertical='bottom', text_rotation=0, \
                                wrap_text=s.get('wr', False), shrink_to_fit=False, indent=0)
                    
            tab = NamedStyle(name=s['name'])
            tab.font = font
            tab.fill = fill
            tab.border = border
            tab.alignment = alignment
            self.wb.add_named_style(tab)
        
    def save(self, pathname='test.xlsx'):
        """Сохраняем документ"""
        self.wb.save(pathname)
        
    def num2col(self, colnumber):
        """Преобразует числовой номер столбца в буквенный"""
        num2col = ""
        while colnumber > 0:
            i = int(colnumber % 26)
            if i > 0:
                num2col = chr(64 + i) + num2col
            else:
                num2col = "Z" + num2col
                colnumber = colnumber - 1
            colnumber = int(colnumber / 26)
        return num2col

    def columnsize(self, clm, sz):
        """Устанавливает размеры столбцов"""

        if type(sz) not in(list, tuple):
            col = self.num2col(sz)
            self.ws.column_dimensions[col].width = sz + 0.7109375
        else:
            for i, cs in enumerate(sz):
                col = self.num2col(clm+i)
                self.ws.column_dimensions[col].width = cs + 0.7109375

    def rowsize(self, rw, sz):
        """Устанавливает размеры строк"""
        if type(sz) != list:
            self.wb.ActiveSheet.Rows(rw).RowHeight = sz
        else:
            for (val, i) in zip(sz, range(len(sz))):
                self.wb.ActiveSheet.Rows(rw+i).RowHeight = val

    def columnfit(self, clm, ln=1):
        """Установка авторазмера колонок"""
        #не адаптированно
        cells = self.num2col(clm) + ":" + self.num2col(clm+ln-1)
        self.wb.ActiveSheet.Columns(cells).EntireColumn.AutoFit

    def rowfit(self, rw, ln=1):
        """Установка авторазмера строк"""
        #не адаптированно
        cells = str(rw) + ":" + str(rw+ln-1)
        self.wb.ActiveSheet.Rows(cells).EntireRow.AutoFit

    def paintcells(self, rw, clm, ln=1, tc=1, ink=2, tas=0):
        """
        Раскрашивание ячеек
        tc   - ThemeColor заливка
        ink - чернила
        tas - TintAndShade номер яркости 0-без затемнения 1-тёмное 5-яркое
        """
        #не адаптированно
        tint1 = {0:0, 1:-0.0499893185216834, 2: -0.149998474074526, 3:-0.249977111117893, 4:-0.349986266670736, 5:-0.499984740745262}
        tint2 = {0:0, 1:0.499984740745262, 2:0.349986266670736, 3:0.249977111117893, 4:0.149998474074526, 5:0.0499893185216834}
        tint3 = {0:0, 1:-0.0999786370433668, 2:-0.249977111117893, 3:-0.499984740745262, 4:-0.749992370372631, 5:-0.899990844447157}
        tint4 = {0:0, 1:0.799981688894314, 2:0.599993896298105, 3:0.399975585192419, 4:-0.249977111117893, 5:-0.499984740745262}
        if tc<1:
            tc = 1
        if tc>10:
            tc = 10
        if tc==1:
            TintAndShade = tint1[tc]
        if tc==2:
            TintAndShade = tint2[tc]
        if tc==3:
            TintAndShade = tint3[tc]
        if tc>3:
            TintAndShade = tint4[tc]
        
        cells = self.num2col(clm)+str(rw)+":"+self.num2col(clm+ln-1)+str(rw)
        
        range = self.wb.ActiveSheet.Range(cells)
        range.Font.ThemeColor = ink
        range.Font.TintAndShade = 0
        range.Interior.Pattern = 1
        range.Interior.PatternColorIndex = -4105
        range.Interior.ThemeColor = tc
        range.Interior.TintAndShade = TintAndShade
        range.Interior.PatternTintAndShade = 0

    def paintcellsnone(self, rw, clm, ln=1):
        """Цвет заливки ячейки НЕТ"""
        #не адаптированно
        cells = self.num2col(clm)+str(rw)+":"+self.num2col(clm+ln-1)+str(rw)
        range = self.wb.ActiveSheet.Range(cells)
        range.Interior.Pattern = -4142
        range.Interior.TintAndShade = 0
        range.Interior.PatternTintAndShade = 0
        range.Font.ColorIndex = -4105
        range.Font.TintAndShade = 0

    def txtformat(self, rw, clm, halign=None, valign=None, wrap='f', nf=[''], \
                  ort=[None], fsz=[None], bold=None, italic='f'):
        """Выравнивание текста в ячейках по горизонтали и вертикали; перенос текста; формат числа
           halign: l-xlLeft, r-xlRight, c-xlCenter
           valign: t-xlTop, c-xlCenter, b-xlBottom
           wrap: f-False, t-True
           nf-NumberFormat: "General" по умолчанию "Общий". Передаётся списком
           ort-orientation: Ориентация текста (поворот в градусах)
           bold - жирный текст f-False, t-True
           """
        al = {'l':'left', 'r':'right', 'c':'center', 't':'top', 'b':'bottom'}
        ft = {'f':False, 't':True}
        if halign:
            halign = halign.lower()
        if valign:
            valign = valign.lower()
        wrap = wrap.lower()
        if bold:
            bold = bold.lower()
        lst = [halign, valign, wrap, nf, ort, fsz, bold]
        rwlen = len(max((i for i in lst if i is not None), key=len))
        for i in range(rwlen):
            cells = self.num2col(clm+i)+str(rw)
            h = al[halign[min(i,len(halign)-1)]] if halign else None
            v = al[valign[min(i,len(valign)-1)]] if valign else None
            w = ft[wrap[min(i,len(wrap)-1)]]
            n = nf[min(i,len(nf)-1)]
            b = ft[bold[min(i,len(bold)-1)]] if bold else None
            it = ft[italic[min(i,len(italic)-1)]]
            o = ort[min(i,len(ort)-1)]
            fz = fsz[min(i,len(fsz)-1)]

            self.ws[cells].alignment = Alignment(horizontal=h, vertical=v, \
                                                 text_rotation=o, \
                                                 wrap_text=w)
            self.ws[cells].font = Font(size=fz, bold=b, italic=it)
            self.ws[cells].number_format = n



    def gridset(self, wt='xlThin', ls='xlContinuous', tc=2):
        """Настройка сетки"""
        self.themecolor = tc
        # Weight - толщины линий
        weight = {'xlThin':2, 'xlHairline':1, 'xlThick':4, 'xlMedium':-4138}
        self.weight = weight[wt]
        # LineStyle - стиль линий
        linestyle = {'xlContinuous':1, 'xlDot':-4118, 'xlDash':-4115, 'xlDashDot':4, 'xlNone':-4142, 'xlDouble':-4119, 'xlSlantDashDot':13}
        self.linestyle = linestyle[ls]

    def grid(self, rw=0, clm=0, ln=0, hg=0, sd='lrud'):
        """Отрисовка сетки"""
        cells = self.num2col(clm)+str(rw)+":"+self.num2col(clm+abs(ln)-1)+str(rw+abs(hg)-1)
        sd = sd.lower()
        side = {'l':7, 'u':8, 'd':9, 'r':10, 'v':11, 'h':12}
        for i in sd:
            gr = self.wb.ActiveSheet.Range(cells).Borders(side[i])
            gr.LineStyle = self.linestyle
            gr.ThemeColor = self.themecolor
            gr.Weight = self.weight
            gr.TintAndShade = 0

    def movesheet(self, s1, s2):
        """Перемещает лист s1 перед s2"""
        self.wb.Sheets(s1).Move(Before=self.wb.Sheets(s2))

    def picinsert(self, rw, clm, ln=1, hg=1, file='', hor='c', ver='c'):
        """Вставка картинки"""
        if not os.path.exists(file):
            return None
        cells = self.num2col(clm)+str(rw)+":"+self.num2col(clm+abs(ln)-1)+str(rw+abs(hg)-1)
        TargetCells = self.wb.ActiveSheet.Range(cells)
        pic = self.wb.ActiveSheet.Shapes.AddPicture(file, 0, 1, -1, -1, -1, -1)
        lWscale = pic.Height/pic.Width
        pic.LockAspectRatio = 0
        pic.Height = TargetCells.Height
        pic.Width  = TargetCells.Width
        Aspect = TargetCells.Height/TargetCells.Width
        if lWscale < Aspect:
            pic.ScaleHeight((lWscale/Aspect), 0, 0)
            pic.ScaleHeight(0.9, 0, 0)
            pic.ScaleWidth(0.9, 0, 0)
        else:
            pic.ScaleWidth((Aspect/lWscale), 0, 0)
        hor = hor.lower()
        ver = ver.lower()
        ph = pic.Height
        pw = pic.Width
        tch = TargetCells.Height
        tcw = TargetCells.Width
        hl = TargetCells.Left
        hc = TargetCells.Left + (tcw - pw)/2
        hr = TargetCells.Left + (tcw - pw)
        vt = TargetCells.Top
        vc = TargetCells.Top + (tch - ph)/2
        vb = TargetCells.Top + (tch - ph)
        halign = {'l':hl, 'c':hc, 'r':hr}
        valign = {'t':vt, 'c':vc, 'b':vb}
        pic.Top = valign[ver]
        pic.Left = halign[hor]
        pic.LockAspectRatio = 1

    def putval(self, row=1, column=1, value=''):
        """Запись данных в ячейки"""
        
        if type(value) in (list, tuple):
            for col, val in enumerate(value):
                self.ws.cell(row, column+col, val)
        else:
            self.ws.cell(row, column, value)
        return row+1

    def printarea(self, rw, clm, ln, hg):
        """Задаёт область печати"""
        cells = self.num2col(clm)+str(rw)+":"+self.num2col(clm+abs(ln)-1)+str(rw+abs(hg)-1)
        self.wb.ActiveSheet.PageSetup.PrintArea = cells

    def table_style(self, rw, clm,ln, hg, tabname, style='TableStyleLight2', st=True):
        """Задать таблицу стилей"""
        cells = self.num2col(clm)+str(rw)+":"+self.num2col(clm+abs(ln)-1)+str(rw+abs(hg)-1)
        range = self.wb.ActiveSheet.Range(cells)
        self.wb.ActiveSheet.ListObjects.Add(1,range,0,1).Name = tabname
        self.wb.ActiveSheet.ListObjects(tabname).TableStyle = style
        self.wb.ActiveSheet.ListObjects(tabname).ShowTotals = st
        range.AutoFilter()
    
    def format_cond(self, rang, tp, op, f1, f2):
        """Условное форматирование
           rang - диапозон ("A:C")
           tp - Type Определяет, основан ли условный формат на значении ячейки или выражении
           op - Operator Условный оператор формата. Может быть одна из следующих констант XlFormatConditionOperator: 
                xlBetween, xlEqual, xlGreater, xlGreaterEqual, xlLess, xlLessEqual, xlNotBetween, или xlNotEqual.
                Если Тип - xlExpression, аргумент Оператора проигнорирован
           f1 - Formula1 Значение или выражение связанное с условным форматированием. Может быть постоянное значение, строковое значение, ссылка на ячейку или формула.
           f2 - Formula2 Значение или выражение связанное со второй частью условного форматирования, когда Оператор - xlBetween или xlNotBetween (иначе, этот параметр проигнорирован).
                Может быть постоянное значение, строковое значение, ссылка на ячейку или формула.
        """
        self.wb.ActiveSheet.Range(rang).FormatConditions.Add(tp, op, f1, f2)

    def styletorange(self, rang, style):
        """Применение стиля к диапазону ячеек"""
        
        cells = self.ws[rang]
        for rng in cells:
            for cell in rng:
                cell.style = style
