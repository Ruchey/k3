# -*- coding: utf-8 -*-

import os
import openpyxl

from openpyxl.styles import PatternFill, Border, Side, Alignment, Font, NamedStyle
from openpyxl.drawing.image import Image
from openpyxl.utils.units import cm_to_EMU, EMU_to_inch, pixels_to_EMU
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.worksheet.datavalidation import DataValidation as dv

__author__ = 'Виноградов А.Г. г.Белгород  август 2015'


def num_to_col(col_number):
    """Преобразует числовой номер столбца в буквенный"""
    num2col = ""
    while col_number > 0:
        i = int(col_number % 26)
        if i > 0:
            num2col = chr(64 + i) + num2col
        else:
            num2col = "Z" + num2col
            col_number = col_number - 1
        col_number = int(col_number / 26)
    return num2col


class Doc:
    """Создание документа Excel
        ExcelDoc - создаёт объект Excel
        newheet - создаёт страницу в книге с заданным именем
        save - сохраняет документ в указанное место
        put_val - добавляет запись в документ
    """

    def __init__(self):
        """Создание объекта Excel, инициализация умолчаний"""
        self.op = openpyxl
        self.dv = dv
        self.wb = self.op.Workbook()
        self.ws = self.wb.active
        self.PORTRAIT = self.ws.ORIENTATION_PORTRAIT
        self.LANDSCAPE = self.ws.ORIENTATION_LANDSCAPE
        self.first = True

        self.sheet_orient = self.PORTRAIT
        self.paper_size = self.ws.PAPERSIZE_A4
        self.right_margin = 0.6
        self.left_margin = 0.6
        self.bottom_margin = 1.6
        self.top_margin = 1.6

        self.theme_color = 2
        self.weight = 2
        self.line_style = 1

        self.center_horizontally = 0
        self.font = "Arial Narrow"
        self.fontsize = 12
        self.display_zeros = False

        # for font Calibri 11
        self.default_row_height = 15
        self.default_row_pixels = 20
        self.default_col_width = 8.43
        self.default_col_pixels = 64
        # Зазор в кол-ве символа добавляемый к ширине столбца
        self.margin = 0.7109375

        self.generate_styles()
        self.F_RUB = openpyxl.styles.numbers.BUILTIN_FORMATS[42]

    def new_sheet(self, name, tab_color=None):
        """Добавление листа"""

        if self.first:
            self.ws.title = name[:31]
            self.first = False
        else:
            self.ws = self.wb.create_sheet(name[:31])

        self.ws.page_setup.orientation = self.sheet_orient
        self.ws.page_setup.paperSize = self.paper_size
        self.ws.page_margins.left = EMU_to_inch(cm_to_EMU(self.left_margin))
        self.ws.page_margins.right = EMU_to_inch(cm_to_EMU(self.right_margin))
        self.ws.page_margins.top = EMU_to_inch(cm_to_EMU(self.top_margin))
        self.ws.page_margins.bottom = EMU_to_inch(cm_to_EMU(self.bottom_margin))
        self.ws.print_options.horizontalCentered = self.center_horizontally
        self.ws.sheet_properties.tabColor = tab_color

    def generate_styles(self):
        """Создание базовых стилей
        Названия стилей:
            'Заголовок 1'
            'Заголовок 2'
            'Заголовок 3'
            'Заголовок 4'
            'Заголовок 5'
            'Заголовок 6'
            'Заголовок 7'
            'Заголовок 8'
            'Итоги 1'
            'Таблица 1'
            'Шапка 1'
        """
        styles = [
            {'name': 'Заголовок 1', 'sc': 'F2F2F2', 'ec': 'F2F2F2', 'bc': 'D9D9D9', 'b': True, 'horAlign': 'right',
             'bb': 'thin', 'ft': 'solid'},
            {'name': 'Заголовок 2', 'sc': 'D9D9D9', 'ec': 'D9D9D9', 'bc': 'BFBFBF', 'b': True, 'horAlign': 'right',
             'bb': 'thin', 'ft': 'solid'},
            {'name': 'Заголовок 3', 'sc': '8DB4E2', 'ec': '8DB4E2', 'bc': '16365C', 'b': True, 'horAlign': 'right',
             'bb': 'thin', 'ft': 'solid'},
            {'name': 'Заголовок 4', 'sc': 'B8CCE4', 'ec': 'B8CCE4', 'bc': '366092', 'b': True, 'horAlign': 'right',
             'bb': 'thin', 'ft': 'solid'},
            {'name': 'Заголовок 5', 'sc': 'E6B8B7', 'ec': 'E6B8B7', 'bc': '963634', 'b': True, 'horAlign': 'right',
             'bb': 'thin', 'ft': 'solid'},
            {'name': 'Заголовок 6', 'sc': 'D8E4BC', 'ec': 'D8E4BC', 'bc': '76933C', 'b': True, 'horAlign': 'right',
             'bb': 'thin', 'ft': 'solid'},
            {'name': 'Заголовок 7', 'sc': 'CCC0DA', 'ec': 'CCC0DA', 'bc': '60497A', 'b': True, 'horAlign': 'right',
             'bb': 'thin', 'ft': 'solid'},
            {'name': 'Заголовок 8', 'sc': 'B7DEE8', 'ec': 'B7DEE8', 'bc': '31869B', 'b': True, 'horAlign': 'right',
             'bb': 'thin', 'ft': 'solid'},
            {'name': 'Итоги 1', 'sc': 'FABF8F', 'ec': 'FABF8F', 'bc': 'E26B0A', 'b': True, 'horAlign': 'right',
             'bb': 'thin', 'bt': 'thick', 'ft': 'solid'},
            {'name': 'Таблица 1', 'bl': 'thin', 'br': 'thin', 'bt': 'thin', 'bb': 'thin', 'wr': True, 'bc': '808080'},
            {'name': 'Шапка 1', 'bl': 'thin', 'br': 'thin', 'bt': 'thin', 'bb': 'thin', 'b': True}
        ]

        for s in styles:
            font = Font(name=self.font, size=self.fontsize, bold=s.get('b', False), italic=s.get('itl', False),
                        vertAlign=s.get('vertAlign'), underline='none', strike=False, color=s.get('txtclr', '000000'))
            fill = PatternFill(fill_type=s.get('ft'), start_color=s.get('sc'), end_color=s.get('ec'))

            border = Border(left=Side(border_style=s.get('bl'), color=s.get('bc', '000000')),
                            right=Side(border_style=s.get('br'), color=s.get('bc', '000000')),
                            top=Side(border_style=s.get('bt'), color=s.get('bc', '000000')),
                            bottom=Side(border_style=s.get('bb'), color=s.get('bc', '000000')),
                            diagonal=Side(border_style=None, color=s.get('bc', '000000')),
                            diagonal_direction=0, outline=Side(border_style=None, color=s.get('bc', '000000')),
                            vertical=Side(border_style=s.get('bv'), color=s.get('bc', '000000')),
                            horizontal=Side(border_style=s.get('bh'), color=s.get('bc', '000000'))
                            )
            alignment = Alignment(horizontal=s.get('horAlign', 'left'),
                                  vertical='bottom', text_rotation=0,
                                  wrap_text=s.get('wr', False), shrink_to_fit=False, indent=0)

            tab = NamedStyle(name=s['name'])
            tab.font = font
            tab.fill = fill
            tab.border = border
            tab.alignment = alignment
            tab.number_format = None
            self.wb.add_named_style(tab)

    def save(self, pathname='test.xlsx'):
        """Сохраняем документ"""
        self.wb.save(pathname)

    def col_size(self, clm, sz):
        """Устанавливает размеры столбцов"""

        if type(sz) not in (list, tuple):
            col = num_to_col(clm)
            self.ws.column_dimensions[col].width = sz + self.margin # 0.7109375
        else:
            for i, val in enumerate(sz):
                col = num_to_col(clm + i)
                self.ws.column_dimensions[col].width = val + self.margin # 0.7109375

    def get_col_size(self, col=1, cnt=1):
        """Возвращает ширину колонки или сумму ширин колонок в единицах документа и пикселях (symbols, px)
        Keyword arguments:
            col -- int номер начального столбца
            cnt -- int количество столбцов
        Returns:
            tuple (symbols, px)
        """
        def _col_px(width):
            max_digit_width = 7  # For Calabri 11.
            padding = 5
            if width is not None:
                if width < 1:
                    pixels = int(width * (max_digit_width + padding) + 0.5)
                else:
                    pixels = int(width * max_digit_width + 0.5) + padding
            else:
                pixels = self.default_col_pixels
            return pixels

        size = 0
        for i in range(cnt):
            clm = num_to_col(col+i)
            sz = self.ws.column_dimensions[clm].width
            if sz is None:
                sz = self.default_col_width
            size += sz
        size = round(size - self.margin, 2)
        px = _col_px(size)
        return size, px

    def row_size(self, rw, sz):
        """Устанавливает размеры строк"""

        if type(sz) not in (list, tuple):
            self.ws.row_dimensions[rw].height = sz
        if type(sz) in (list, tuple):
            for i, val in enumerate(sz):
                self.ws.row_dimensions[rw + i].height = val

    def get_row_size(self, rw=1, cnt=1):
        """Возвращает высоту строки сумму ширин высот в единицах документа и пикселях (symbols, px)
        Keyword arguments:
            rw -- int номер начальной строки
            cnt -- int количество строк
        Returns:
            tuple (symbols, px)
        """
        def _row_px(height):
            if height is not None:
                pixels = int(4.0 / 3.0 * height)
            else:
                pixels = int(4.0 / 3.0 * self.default_row_height)
            return pixels

        size = 0
        px = 0
        for i in range(cnt):
            sz = self.ws.row_dimensions[rw+i].height
            px += _row_px(sz)
            if sz is None:
                sz = self.default_row_height
            size += sz
        return size, px

    def column_fit(self, clm, ln=1):
        """Установка авторазмера колонок"""
        # TODO: адаптировать под новый модуль
        cells = num_to_col(clm) + ":" + num_to_col(clm + ln - 1)
        self.wb.ActiveSheet.Columns(cells).EntireColumn.AutoFit

    def row_fit(self, rw, ln=1):
        """Установка авторазмера строк"""
        # TODO: адаптировать под новый модуль
        cells = str(rw) + ":" + str(rw + ln - 1)
        self.wb.ActiveSheet.Rows(cells).EntireRow.AutoFit

    def paint_cells(self, range_cells, ink=None, fill=None):
        """Раскрашивание ячеек
        cells - диапазон ячеек формата 'A1:B3'
        ink - цвет шрифта в формате FFF000
        paper - цвет заливки в формате FFF000
        """
        cells = self.ws[range_cells]
        filling = PatternFill(fill_type='solid', start_color=fill, end_color=fill)
        font_color = Font(color=ink)
        if type(cells) is tuple:
            for cell in cells:
                for i in cell:
                    if ink:
                        i.font = font_color
                    if fill:
                        i.fill = filling
        else:
            if ink:
                cells.font = font_color
            if fill:
                cells.fill = filling

    def formatting(self, rw=1, col=1, **kwargs):
        """Форматирует ячейки
        Keyword arguments:
            rw -- int Строка
            col -- int - Колонка
            ha -- str horizontal align = l - xlLeft, r - xlRight, c - xlCenter
            va -- str vertical align =  t - xlTop, c - xlCenter, b - xlBottom
            wrap -- str f - False, t - True
            nf -- list|tuple|str NumberFormat "General" по умолчанию "Общий". Передаётся списком
            rot -- list|tuple|int rotate Ориентация текста (поворот в градусах)
            bld -- str bold жирный текст f-False, t-True
            sz -- list|tuple|int size font Размер шрифта
            itl -- str italic Курсивный шрифт
        """
        ha = kwargs.get('ha', None)
        va = kwargs.get('va', None)
        wrap = kwargs.get('wrap', None)
        bld = kwargs.get('bld', None)
        itl = kwargs.get('itl', None)
        nf = kwargs.get('nf', None)
        rot = kwargs.get('rot', None)
        sz = kwargs.get('sz', None)

        align = {'l': 'left', 'r': 'right', 'c': 'center', 't': 'top', 'bld': 'bottom'}
        ft = {'f': False, 't': True}
        if ha:
            ha = ha.lower()
        if va:
            va = va.lower()
        if wrap:
            wrap = wrap.lower()
        if bld:
            bld = bld.lower()
        if type(rot) not in (list, tuple):
            rot = (rot,)
        if type(sz) not in [list, tuple]:
            sz = (sz,)
        if type(nf) not in [list, tuple]:
            nf = (nf,)
        lst = [ha, va, wrap, bld, itl, nf, rot, sz]
        rw_len = len(max((i for i in lst if i is not None), key=len))
        for i in range(rw_len):
            cells = num_to_col(col + i) + str(rw)
            h = align[ha[min(i, len(ha) - 1)]] if ha else None
            v = align[va[min(i, len(va) - 1)]] if va else None
            w = ft[wrap[min(i, len(wrap) - 1)]] if wrap else None
            b = ft[bld[min(i, len(bld) - 1)]] if bld else self.ws[cells].font.b
            it = ft[itl[min(i, len(itl) - 1)]] if itl else self.ws[cells].font.i
            r = rot[min(i, len(rot) - 1)] if any(rot) else self.ws[cells].alignment.textRotation
            fz = sz[min(i, len(sz) - 1)] if any(sz) else self.ws[cells].font.sz
            n = nf[min(i, len(nf) - 1)] if any(nf) else self.ws[cells].number_format

            self.ws[cells].alignment = Alignment(horizontal=h, vertical=v,
                                                 text_rotation=r,
                                                 wrap_text=w)
            self.ws[cells].font = Font(size=fz, bold=b, italic=it)
            self.ws[cells].number_format = n

    def grid_set(self, wt='xlThin', ls='xlContinuous', tc=2):
        """Настройка сетки"""
        self.theme_color = tc
        # Weight - толщины линий
        weight = {'xlThin': 2, 'xlHairline': 1, 'xlThick': 4, 'xlMedium': -4138}
        self.weight = weight[wt]
        # LineStyle - стиль линий
        line_style = {'xlContinuous': 1, 'xlDot': -4118, 'xlDash': -4115, 'xlDashDot': 4, 'xlNone': -4142,
                      'xlDouble': -4119, 'xlSlantDashDot': 13}
        self.line_style = line_style[ls]

    def grid(self, rw=0, clm=0, ln=0, hg=0, sd='lrud'):
        """Отрисовка сетки"""
        # TODO: адаптировать под новый модуль
        cells = num_to_col(clm) + str(rw) + ":" + num_to_col(clm + abs(ln) - 1) + str(rw + abs(hg) - 1)
        sd = sd.lower()
        side = {'l': 7, 'u': 8, 'd': 9, 'r': 10, 'v': 11, 'h': 12}
        for i in sd:
            gr = self.wb.ActiveSheet.Range(cells).Borders(side[i])
            gr.LineStyle = self.line_style
            gr.ThemeColor = self.theme_color
            gr.Weight = self.weight
            gr.TintAndShade = 0

    def move_sheet(self, s1, s2):
        """Перемещает лист s1 перед s2"""
        # TODO: адаптировать под новый модуль
        self.wb.Sheets(s1).Move(Before=self.wb.Sheets(s2))

    def pic_insert(self, rw=1, col=1, **kwargs):
        """Вставка картинки
        Keyword arguments:
            rw -- int номер строки
            col -- int номер колонки
            max_col -- int количество занимаемых колонок
            max_row -- int количество занимаемых строк
            path -- str путь к картинке
            align -- str буквенное обозначеение горизонтального выравнивания картинки:
                l - left
                r - right
                c - center
            valign -- str буквенное обозначеение вертикального выравнивания картинки:
                t - top
                c - center
                b - bottom
            cf -- float коэффициент размера картинки, 1 = 100%
            fit -- bool вписать картинку в пределы ячеек default(True)
            px -- int количество пикселей зазор до границы
        """

        path = kwargs.get('path', '')
        cf = kwargs.get('cf', 1)
        max_col = kwargs.get('max_col', 1)
        max_row = kwargs.get('max_row', 1)
        align = kwargs.get('align', 'l')
        valign = kwargs.get('valign', 't')
        fit = kwargs.get('fit', True)
        px = kwargs.get('px', 0)

        if not os.path.exists(path):
            return None

        img = Image(path)
        img_h, img_w = img.height, img.width
        p2e = pixels_to_EMU

        cell_h = self.get_row_size(rw, max_row)[1]
        cell_w = self.get_col_size(col, max_col)[1]
        if fit:
            img_ratio = img_w / img_h
            img_h = cell_h
            img_w = img_h * img_ratio
            if img_w > cell_w:
                img_w = cell_w
                img_h = img_w / img_ratio
        img_h *= cf
        img_w *= cf
        img_size = XDRPositiveSize2D(p2e(img_w), p2e(img_h))

        d_h = cell_h - img_h
        d_w = cell_w - img_w
        hl = px
        hr = d_w - px
        hc = cell_w / 2 - img_w / 2
        vt = px
        vb = d_h - px
        vc = d_h / 2
        h_align = {'l': hl, 'c': hc, 'r': hr}
        v_align = {'t': vt, 'c': vc, 'b': vb}
        row_offset = p2e(v_align[valign])
        col_offset = p2e(h_align[align])

        marker = AnchorMarker(col=col-1, colOff=col_offset, row=rw-1, rowOff=row_offset)
        img.anchor = OneCellAnchor(_from=marker, ext=img_size)
        self.ws.add_image(img)

    def put_val(self, row=1, col=1, value=''):
        """Запись данных в ячейки
        Аргументы:
            row -- int Строка, начиная с 1
            col -- int Столбец, начиная с 1
        """
        if type(value) in (list, tuple):
            for i, val in enumerate(value):
                self.ws.cell(row, col+i, val)
        else:
            self.ws.cell(row, col, value)
        return row + 1

    def print_area(self, rw, clm, ln, hg):
        """Задаёт область печати"""
        # TODO: адаптировать под новый модуль
        cells = num_to_col(clm) + str(rw) + ":" + num_to_col(clm + abs(ln) - 1) + str(rw + abs(hg) - 1)
        self.wb.ActiveSheet.PageSetup.PrintArea = cells

    def format_cond(self, rang, tp, op, f1, f2):
        """Условное форматирование
           rang - диапозон ("A:C")
           tp - Type Определяет, основан ли условный формат на значении ячейки или выражении
           op - Operator Условный оператор формата. Может быть одна из следующих констант XlFormatConditionOperator:
                xlBetween, xlEqual, xlGreater, xlGreaterEqual, xlLess, xlLessEqual, xlNotBetween, или xlNotEqual.
                Если Тип - xlExpression, аргумент Оператора проигнорирован
           f1 - Formula1 Значение или выражение связанное с условным форматированием. Может быть постоянное значение, строковое значение, ссылка на ячейку или формула.
           f2 - Formula2 Значение или выражение связанное со второй частью условного форматирования, когда Оператор - xlBetween или xlNotBetween (иначе, этот параметр проигнорирован).
                Может быть постоянное значение, строковое значение, ссылка на ячейку или формула.
        """
        # TODO: адаптировать под новый модуль
        self.wb.ActiveSheet.Range(rang).FormatConditions.Add(tp, op, f1, f2)

    def style_to_range(self, rang, style):
        """Применение стиля к диапазону ячеек
        Аргументы:
            rang -- str 'A1:F6'
            style -- str:
                'Заголовок 1'
                'Заголовок 2'
                'Заголовок 3'
                'Заголовок 4'
                'Заголовок 5'
                'Заголовок 6'
                'Заголовок 7'
                'Заголовок 8'
                'Итоги 1'
                'Таблица 1'
                'Шапка 1'
        """
        cells = self.ws[rang]
        for rng in cells:
            for cell in rng:
                cell.style = style

    def named_ranges(self, name, attr_txt):
        """Создаёт именованный диапазон
        Аргументы:
            name -- str например, 'res_sum'
            attr_txt -- str например, 'sheet!$H$10'
        """
        new_range = openpyxl.workbook.defined_name.DefinedName(name, attr_text=attr_txt)
        self.wb.defined_names.append(new_range)
