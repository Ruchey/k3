# -*- coding: utf-8 -*-
__author__ = 'Виноградов А.Г. г.Белгород  август 2015'

import openpyxl
import os

from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, NamedStyle
from openpyxl.utils.units import cm_to_EMU, EMU_to_inch, pixels_to_points


def num_to_col(col_number):
    """Преобразует числовой номер столбца в буквенный"""
    num2col = ""
    while col_number > 0:
        i = int(col_number % 26)
        if i > 0:
            num2col = chr(64 + i) + num2col
        else:
            num2col = "Z" + num2col
            col_number = col_number - 1
        col_number = int(col_number / 26)
    return num2col


class Doc:
    """Создание документа Excel
        ExcelDoc - создаёт объект Excel
        newheet - создаёт страницу в книге с заданным именем
        save - сохраняет документ в указанное место
        put_val - добавляет запись в документ
    """

    def __init__(self):
        """Создание объекта Excel, инициализация умолчаний"""
        self.xl = openpyxl
        self.wb = self.xl.Workbook()
        self.ws = self.wb.active
        self.PORTRAIT = self.ws.ORIENTATION_PORTRAIT
        self.LANDSCAPE = self.ws.ORIENTATION_LANDSCAPE
        self.first = True

        self.sheet_orient = self.PORTRAIT
        self.paper_size = self.ws.PAPERSIZE_A4
        self.right_margin = 0.6
        self.left_margin = 0.6
        self.bottom_margin = 1.6
        self.top_margin = 1.6

        self.theme_color = 2
        self.weight = 2
        self.line_style = 1

        self.center_horizontally = 0
        self.font = "Arial Narrow"
        self.fontsize = 12
        self.display_zeros = False

        self.generate_styles()
        self.F_RUB = openpyxl.styles.numbers.BUILTIN_FORMATS[42]

    def new_sheet(self, name, tab_color=None):
        """Добавление листа"""

        if self.first:
            self.ws.title = name[:31]
            self.first = False
        else:
            self.ws = self.wb.create_sheet(name[:31])

        self.ws.page_setup.orientation = self.sheet_orient
        self.ws.page_setup.paperSize = self.paper_size
        self.ws.page_margins.left = EMU_to_inch(cm_to_EMU(self.left_margin))
        self.ws.page_margins.right = EMU_to_inch(cm_to_EMU(self.right_margin))
        self.ws.page_margins.top = EMU_to_inch(cm_to_EMU(self.top_margin))
        self.ws.page_margins.bottom = EMU_to_inch(cm_to_EMU(self.bottom_margin))
        self.ws.print_options.horizontalCentered = self.center_horizontally
        self.ws.sheet_properties.tabColor = tab_color

    def generate_styles(self):
        """Создание базовых стилей"""
        styles = [
            {'name': 'Заголовок 1', 'sc': 'F2F2F2', 'ec': 'F2F2F2', 'bc': 'D9D9D9', 'bold': True, 'horAlign': 'right',
             'bb': 'thin', 'ft': 'solid'},
            {'name': 'Заголовок 2', 'sc': 'D9D9D9', 'ec': 'D9D9D9', 'bc': 'BFBFBF', 'bold': True, 'horAlign': 'right',
             'bb': 'thin', 'ft': 'solid'},
            {'name': 'Заголовок 3', 'sc': '8DB4E2', 'ec': '8DB4E2', 'bc': '16365C', 'bold': True, 'horAlign': 'right',
             'bb': 'thin', 'ft': 'solid'},
            {'name': 'Заголовок 4', 'sc': 'B8CCE4', 'ec': 'B8CCE4', 'bc': '366092', 'bold': True, 'horAlign': 'right',
             'bb': 'thin', 'ft': 'solid'},
            {'name': 'Заголовок 5', 'sc': 'E6B8B7', 'ec': 'E6B8B7', 'bc': '963634', 'bold': True, 'horAlign': 'right',
             'bb': 'thin', 'ft': 'solid'},
            {'name': 'Заголовок 6', 'sc': 'D8E4BC', 'ec': 'D8E4BC', 'bc': '76933C', 'bold': True, 'horAlign': 'right',
             'bb': 'thin', 'ft': 'solid'},
            {'name': 'Заголовок 7', 'sc': 'CCC0DA', 'ec': 'CCC0DA', 'bc': '60497A', 'bold': True, 'horAlign': 'right',
             'bb': 'thin', 'ft': 'solid'},
            {'name': 'Заголовок 8', 'sc': 'B7DEE8', 'ec': 'B7DEE8', 'bc': '31869B', 'bold': True, 'horAlign': 'right',
             'bb': 'thin', 'ft': 'solid'},
            {'name': 'Таблица 1', 'bl': 'thin', 'br': 'thin', 'bt': 'thin', 'bb': 'thin', 'wr': True, 'bc': '808080'},
            {'name': 'Шапка 1', 'bl': 'thin', 'br': 'thin', 'bt': 'thin', 'bb': 'thin', 'bold': True}
        ]

        for s in styles:
            font = Font(name=self.font, size=self.fontsize, bold=s.get('bold', False), italic=s.get('italic', False),
                        vertAlign=s.get('vertAlign'), underline='none', strike=False, color=s.get('txtclr', '000000'))
            fill = PatternFill(fill_type=s.get('ft'), start_color=s.get('sc'), end_color=s.get('ec'))

            border = Border(left=Side(border_style=s.get('bl'), color=s.get('bc', '000000')),
                            right=Side(border_style=s.get('br'), color=s.get('bc', '000000')),
                            top=Side(border_style=s.get('bt'), color=s.get('bc', '000000')),
                            bottom=Side(border_style=s.get('bb'), color=s.get('bc', '000000')),
                            diagonal=Side(border_style=None, color=s.get('bc', '000000')),
                            diagonal_direction=0, outline=Side(border_style=None, color=s.get('bc', '000000')),
                            vertical=Side(border_style=s.get('bv'), color=s.get('bc', '000000')),
                            horizontal=Side(border_style=s.get('bh'), color=s.get('bc', '000000'))
                            )
            alignment = Alignment(horizontal=s.get('horAlign', 'left'),
                                  vertical='bottom', text_rotation=0,
                                  wrap_text=s.get('wr', False), shrink_to_fit=False, indent=0)

            tab = NamedStyle(name=s['name'])
            tab.font = font
            tab.fill = fill
            tab.border = border
            tab.alignment = alignment
            self.wb.add_named_style(tab)

    def save(self, pathname='test.xlsx'):
        """Сохраняем документ"""
        self.wb.save(pathname)

    def column_size(self, clm, sz):
        """Устанавливает размеры столбцов"""

        if type(sz) not in (list, tuple):
            col = num_to_col(sz)
            self.ws.column_dimensions[col].width = sz + 0.7109375
        else:
            for i, cs in enumerate(sz):
                col = num_to_col(clm + i)
                self.ws.column_dimensions[col].width = cs + 0.7109375

    def row_size(self, rw, sz):
        """Устанавливает размеры строк"""
        if type(sz) != list:
            self.wb.ActiveSheet.Rows(rw).RowHeight = sz
        else:
            for (val, i) in zip(sz, range(len(sz))):
                self.wb.ActiveSheet.Rows(rw + i).RowHeight = val

    def column_fit(self, clm, ln=1):
        """Установка авторазмера колонок"""
        # TODO: адаптировать под новый модуль
        cells = num_to_col(clm) + ":" + num_to_col(clm + ln - 1)
        self.wb.ActiveSheet.Columns(cells).EntireColumn.AutoFit

    def row_fit(self, rw, ln=1):
        """Установка авторазмера строк"""
        # TODO: адаптировать под новый модуль
        cells = str(rw) + ":" + str(rw + ln - 1)
        self.wb.ActiveSheet.Rows(cells).EntireRow.AutoFit

    def paint_cells(self, range_cells, ink=None, fill=None):
        """Раскрашивание ячеек
        cells - диапазон ячеек формата 'A1:B3'
        ink - цвет шрифта в формате FFF000
        paper - цвет заливки в формате FFF000
        """
        cells = self.ws[range_cells]
        filling = PatternFill(fill_type='solid', start_color=fill, end_color=fill)
        font_color = Font(color=ink)
        if type(cells) is tuple:
            for cell in cells:
                for i in cell:
                    if ink:
                        i.font = font_color
                    if fill:
                        i.fill = filling
        else:
            if ink:
                cells.font = font_color
            if fill:
                cells.fill = filling

    def paint_cells_none(self, rw, clm, ln=1):
        """Цвет заливки ячейки НЕТ"""
        # TODO: адаптировать под новый модуль
        cells = num_to_col(clm) + str(rw) + ":" + num_to_col(clm + ln - 1) + str(rw)
        range = self.wb.ActiveSheet.Range(cells)
        range.Interior.Pattern = -4142
        range.Interior.TintAndShade = 0
        range.Interior.PatternTintAndShade = 0
        range.Font.ColorIndex = -4105
        range.Font.TintAndShade = 0

    def txt_format(self, rw, clm, h_align=None, v_align=None, wrap='f', nf=('',),
                   ort=(None,), fsz=(None,), bold=None, italic='f'):
        """Выравнивание текста в ячейках по горизонтали и вертикали; перенос текста; формат числа
           h_align: l-xlLeft, r-xlRight, c-xlCenter
           v_align: t-xlTop, c-xlCenter, b-xlBottom
           wrap: f-False, t-True
           nf-NumberFormat: "General" по умолчанию "Общий". Передаётся списком
           ort-orientation: Ориентация текста (поворот в градусах)
           bold - жирный текст f-False, t-True
           """
        al = {'l': 'left', 'r': 'right', 'c': 'center', 't': 'top', 'b': 'bottom'}
        ft = {'f': False, 't': True}
        if h_align:
            h_align = h_align.lower()
        if v_align:
            v_align = v_align.lower()
        wrap = wrap.lower()
        if bold:
            bold = bold.lower()
        lst = [h_align, v_align, wrap, nf, ort, fsz, bold]
        rw_len = len(max((i for i in lst if i is not None), key=len))
        for i in range(rw_len):
            cells = num_to_col(clm + i) + str(rw)
            h = al[h_align[min(i, len(h_align) - 1)]] if h_align else None
            v = al[v_align[min(i, len(v_align) - 1)]] if v_align else None
            w = ft[wrap[min(i, len(wrap) - 1)]]
            n = nf[min(i, len(nf) - 1)]
            b = ft[bold[min(i, len(bold) - 1)]] if bold else None
            it = ft[italic[min(i, len(italic) - 1)]]
            o = ort[min(i, len(ort) - 1)]
            fz = fsz[min(i, len(fsz) - 1)]

            self.ws[cells].alignment = Alignment(horizontal=h, vertical=v,
                                                 text_rotation=o,
                                                 wrap_text=w)
            self.ws[cells].font = Font(size=fz, bold=b, italic=it)
            self.ws[cells].number_format = n

    def grid_set(self, wt='xlThin', ls='xlContinuous', tc=2):
        """Настройка сетки"""
        self.theme_color = tc
        # Weight - толщины линий
        weight = {'xlThin': 2, 'xlHairline': 1, 'xlThick': 4, 'xlMedium': -4138}
        self.weight = weight[wt]
        # LineStyle - стиль линий
        line_style = {'xlContinuous': 1, 'xlDot': -4118, 'xlDash': -4115, 'xlDashDot': 4, 'xlNone': -4142,
                      'xlDouble': -4119, 'xlSlantDashDot': 13}
        self.line_style = line_style[ls]

    def grid(self, rw=0, clm=0, ln=0, hg=0, sd='lrud'):
        """Отрисовка сетки"""
        # TODO: адаптировать под новый модуль
        cells = num_to_col(clm) + str(rw) + ":" + num_to_col(clm + abs(ln) - 1) + str(rw + abs(hg) - 1)
        sd = sd.lower()
        side = {'l': 7, 'u': 8, 'd': 9, 'r': 10, 'v': 11, 'h': 12}
        for i in sd:
            gr = self.wb.ActiveSheet.Range(cells).Borders(side[i])
            gr.LineStyle = self.line_style
            gr.ThemeColor = self.theme_color
            gr.Weight = self.weight
            gr.TintAndShade = 0

    def move_sheet(self, s1, s2):
        """Перемещает лист s1 перед s2"""
        # TODO: адаптировать под новый модуль
        self.wb.Sheets(s1).Move(Before=self.wb.Sheets(s2))

    def pic_insert(self, rw, clm, ln=1, hg=1, file='', hor='c', ver='c'):
        """Вставка картинки"""
        # TODO: адаптировать под новый модуль
        if not os.path.exists(file):
            return None
        cells = num_to_col(clm) + str(rw) + ":" + num_to_col(clm + abs(ln) - 1) + str(rw + abs(hg) - 1)
        TargetCells = self.wb.ActiveSheet.Range(cells)
        pic = self.wb.ActiveSheet.Shapes.AddPicture(file, 0, 1, -1, -1, -1, -1)
        lWscale = pic.Height / pic.Width
        pic.LockAspectRatio = 0
        pic.Height = TargetCells.Height
        pic.Width = TargetCells.Width
        Aspect = TargetCells.Height / TargetCells.Width
        if lWscale < Aspect:
            pic.ScaleHeight((lWscale / Aspect), 0, 0)
            pic.ScaleHeight(0.9, 0, 0)
            pic.ScaleWidth(0.9, 0, 0)
        else:
            pic.ScaleWidth((Aspect / lWscale), 0, 0)
        hor = hor.lower()
        ver = ver.lower()
        ph = pic.Height
        pw = pic.Width
        tch = TargetCells.Height
        tcw = TargetCells.Width
        hl = TargetCells.Left
        hc = TargetCells.Left + (tcw - pw) / 2
        hr = TargetCells.Left + (tcw - pw)
        vt = TargetCells.Top
        vc = TargetCells.Top + (tch - ph) / 2
        vb = TargetCells.Top + (tch - ph)
        h_align = {'l': hl, 'c': hc, 'r': hr}
        v_align = {'t': vt, 'c': vc, 'b': vb}
        pic.Top = v_align[ver]
        pic.Left = h_align[hor]
        pic.LockAspectRatio = 1

    def put_val(self, row=1, column=1, value=''):
        """Запись данных в ячейки"""

        if type(value) in (list, tuple):
            for col, val in enumerate(value):
                self.ws.cell(row, column + col, val)
        else:
            self.ws.cell(row, column, value)
        return row + 1

    def print_area(self, rw, clm, ln, hg):
        """Задаёт область печати"""
        # TODO: адаптировать под новый модуль
        cells = num_to_col(clm) + str(rw) + ":" + num_to_col(clm + abs(ln) - 1) + str(rw + abs(hg) - 1)
        self.wb.ActiveSheet.PageSetup.PrintArea = cells

    def table_style(self, rw, clm, ln, hg, tabname, style='TableStyleLight2', st=True):
        """Задать таблицу стилей"""
        # TODO: адаптировать под новый модуль
        cells = num_to_col(clm) + str(rw) + ":" + num_to_col(clm + abs(ln) - 1) + str(rw + abs(hg) - 1)
        range = self.wb.ActiveSheet.Range(cells)
        self.wb.ActiveSheet.ListObjects.Add(1, range, 0, 1).Name = tabname
        self.wb.ActiveSheet.ListObjects(tabname).TableStyle = style
        self.wb.ActiveSheet.ListObjects(tabname).ShowTotals = st
        range.AutoFilter()

    def format_cond(self, rang, tp, op, f1, f2):
        """Условное форматирование
           rang - диапозон ("A:C")
           tp - Type Определяет, основан ли условный формат на значении ячейки или выражении
           op - Operator Условный оператор формата. Может быть одна из следующих констант XlFormatConditionOperator: 
                xlBetween, xlEqual, xlGreater, xlGreaterEqual, xlLess, xlLessEqual, xlNotBetween, или xlNotEqual.
                Если Тип - xlExpression, аргумент Оператора проигнорирован
           f1 - Formula1 Значение или выражение связанное с условным форматированием. Может быть постоянное значение, строковое значение, ссылка на ячейку или формула.
           f2 - Formula2 Значение или выражение связанное со второй частью условного форматирования, когда Оператор - xlBetween или xlNotBetween (иначе, этот параметр проигнорирован).
                Может быть постоянное значение, строковое значение, ссылка на ячейку или формула.
        """
        # TODO: адаптировать под новый модуль
        self.wb.ActiveSheet.Range(rang).FormatConditions.Add(tp, op, f1, f2)

    def style_to_range(self, rang, style):
        """Применение стиля к диапазону ячеек"""

        cells = self.ws[rang]
        for rng in cells:
            for cell in rng:
                cell.style = style
