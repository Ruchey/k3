# -*- coding: utf-8 -*-

import k3Report
import os
import openpyxl
from collections import namedtuple


class Report:
    
    def __init__(self, db, phone='', onelist=1):
        
        self.phone = phone
        self.onelist = onelist
        self.nm = k3Report.nomenclature.Nomenclature(db)
        self.bs = k3Report.base.Base(db)
        self.pn = k3Report.panel.Panel(db)
        # Создаём объект документа
        self.xl = k3Report.doc.DocOpenpyxl()

    def newsheet(self, name):
        'Создаём новый лист с именем'
    
        self.xl.sheetorient = self.xl.PORTRAIT
        self.xl.rightmargin = 0.6
        self.xl.leftmargin = 0.6
        self.xl.bottommargin = 1.6
        self.xl.topmargin = 1.6
        self.xl.centerhorizontally = 0
        self.xl.displayzeros = False
        self.xl.newsheet(name)
        return
    
    def preparation(self):
        """Подготовка бланка"""
    
        to = self.bs.torderinfo()
        main_customer = to.executor
        main_phone = self.phone
        order_number = to.ordernumber     # номер заказа
        order_customer = to.customer      # заказчик
        order_phone = to.telephonenumber  # телефон заказчика
        order_expirationdata = to.orderexpirationdata  # Дата отгрузки
        row = 1
        val1 = ('Заказчик:', main_customer, '', 'Заказ №', order_number)
        val2 = ('Телефон:', main_phone, '', 'Заказчик:', order_customer)
        val3 = ('Дата отгрузки:', order_expirationdata)
        self.xl.txtformat(row, 2, fsz=[12], bold='tfftf', italic='ftfft')
        row = self.xl.putval(row, 2, val1)
        self.xl.txtformat(row, 2, fsz=[12], bold='tfftf', italic='ftfft')
        row = self.xl.putval(row, 2, val2)
        self.xl.txtformat(row, 2, fsz=[12], bold='tfftf', italic='ftfft')
        row = self.xl.putval(row, 2, val3)
        row += 1

        cs = [5, 16, 3, 14, 9, 9, 5, 3, 3, 3, 3, 16]
        self.xl.columnsize(1, cs)
        return row
    
    def totalmat(self, matid, tpp=None):
        """Получаем общее кол-во материала"""
        
        Mat = namedtuple('Mat', 'name unit cnt')
        mat_name = self.nm.property_name(matid, 'Name')
        mat_unit = self.nm.property_name(matid, 'UnitsName')
        mat_cnt = self.nm.matcount(matid, tpp)
        mat = Mat(mat_name, mat_unit, mat_cnt)
        list_bands = self.pn.total_bands_to_panels(self.pn.list_panels(matid, tpp))
        bands = []
        for band in list_bands:
            Band = namedtuple('Band', 'name unit length')
            prop = self.nm.properties(band.priceid)
            band_name = prop.name
            band_unit = prop.unitsname
            band_length = band.length
            bands.append(Band(band_name, band_unit, band_length))
        return (mat, bands)
    
    def list_pan(self, matid, tpp=None):
        """Получаем список деталей панелей"""
    
        idPans = self.pn.list_panels(matid, tpp)
        listPan = []
        keys = ('name', 'thickness', 'article', 'length', 'width', 'cnt',
                'band_x1', 'band_x2', 'band_y1', 'band_y2', 'cpos', 'note')
        for idpan in idPans:
            Pans = namedtuple('Pans', keys)
            if self.pn.form(idpan) == 0:
                telems = self.bs.telems(idpan)
                matprop = self.nm.properties(matid)
                name = telems.name
                thickness = matprop.thickness
                article = matprop.article
                length = self.pn.planelength(idpan)
                width = self.pn.planewidth(idpan)
                cnt = telems.count
                band_x1 = self.pn.band_x1(idpan).thickband
                band_x2 = self.pn.band_x2(idpan).thickband
                band_y1 = self.pn.band_y1(idpan).thickband
                band_y2 = self.pn.band_y2(idpan).thickband
                cpos = telems.commonpos
                pdir = self.pn.dir(idpan)
                note_curvepath = 'Фигурная' if self.pn.curvepath(idpan) else ''
                slotx = list(map('{0.beg}ш{0.width}г{0.depth}'.format,
                                 self.pn.slots_x_par(idpan)))
                sloty = list(map('{0.beg}ш{0.width}г{0.depth}'.format,
                                 self.pn.slots_y_par(idpan)))
                if (45<pdir<=135) or (225<pdir<=315):
                    slotx, sloty = sloty, slotx
                note_slotx = "Паз{0} по X {1}".format("ы" if len(slotx) > 1 else "",
                                                      "; ".join(slotx)) if slotx else ""
                note_sloty = "Паз{0} по Y {1}".format("ы" if len(sloty) > 1 else "",
                                                      "; ".join(sloty)) if sloty else ""
                notes = [note_curvepath, note_slotx, note_sloty]
                note = ". ".join(list(filter(None, notes)))
                pans = Pans(name, thickness, article, length, width, cnt, band_x1,
                            band_x2, band_y1, band_y2, cpos, note
                            )
                listPan.append(pans)
        newlist = k3Report.utils.groupbykey(listPan, 'cpos', 'cnt')
        return newlist
    
    def rep_pan(self, name, mat, tpp=None):
        """Создаём лист с данными"""
        self.newsheet(name)
        row = self.preparation()
        rowstart = row

        for imat in mat:
            mt, bands = self.totalmat(imat.id, tpp)
            if mt:
                val = (mt.name, "", "", mt.unit, mt.cnt)
                row = self.xl.putval(row, 2, val)
                self.xl.ws.merge_cells('B{0}:D{0}'.format(row-1))
            if bands:
                for band in bands:
                    val = (band.name, "", "", band.unit, band.length)
                    row = self.xl.putval(row, 2, val)
                    self.xl.ws.merge_cells('B{0}:D{0}'.format(row-1))
        rang = 'B{0}:F{1}'.format(rowstart, row-1)
        self.xl.styletorange(rang, 'Таблица 1')
        row += 1
        row1 = row
        # Формируем шапку
        val = ('№дет', 'Название', 'толщ', 'Артикул', 'Длина', 'Ширина', \
               'шт', 'Кромка', '', '', '', 'Примечание')
        row = self.xl.putval(row, 1, val)
        val = ('Длина', '', 'Ширина')
        row = self.xl.putval(row, 8, val)
        row2 = row
        val = (1, 3, 2, 4)
        row = self.xl.putval(row, 8, val)
        self.xl.ws.merge_cells('A{0}:A{1}'.format(row1, row2))
        self.xl.ws.merge_cells('B{0}:B{1}'.format(row1, row2))
        self.xl.ws.merge_cells('C{0}:C{1}'.format(row1, row2))
        self.xl.ws.merge_cells('D{0}:D{1}'.format(row1, row2))
        self.xl.ws.merge_cells('E{0}:E{1}'.format(row1, row2))
        self.xl.ws.merge_cells('F{0}:F{1}'.format(row1, row2))
        self.xl.ws.merge_cells('G{0}:G{1}'.format(row1, row2))
        self.xl.ws.merge_cells('H{0}:K{0}'.format(row1))
        self.xl.ws.merge_cells('H{0}:I{0}'.format(row1+1))
        self.xl.ws.merge_cells('J{0}:K{0}'.format(row1+1))
        self.xl.ws.merge_cells('L{0}:L{1}'.format(row1, row2))
        self.xl.styletorange('A{0}:L{1}'.format(row1, row), 'Шапка 1')
        self.xl.txtformat(row1, 1, halign='c', valign='cccccccccccc')
        self.xl.txtformat(row1, 3, ort=[90,])
        for imat in mat:
            pans = self.list_pan(imat.id, tpp)
            rang = 'A{0}:L{0}'.format(row)
            self.xl.styletorange(rang, 'Заголовок 1')
            self.xl.ws.merge_cells(rang)
            row = self.xl.putval(row, 1, imat.name)
            rowstart = row
            for p in pans:
                val = (p.cpos, p.name, imat.thickness, imat.article, \
                       p.length, p.width, p.cnt, \
                       p.band_x1, p.band_x2, p.band_y1, p.band_y2, p.note
                       )
                row = self.xl.putval(row, 1, val)
            rang = 'A{0}:L{1}'.format(rowstart, row-1)
            self.xl.styletorange(rang, 'Таблица 1')

    def makereport(self, tpp=None):
        # 'Формируем данные'
        matid = self.nm.matbyuid(2, tpp)
        if not matid:
            return
        listmat = []
        listDSP = []
        listMDF = []
        listDVP = []
        listGLS = []
        allmat = []
        for i in matid:
            prop = self.nm.properties(i)
            if prop.mattypeid == 128:
                listDSP.append(prop)
            elif prop.mattypeid == 64:
                listMDF.append(prop)
            elif prop.mattypeid == 37:
                listDVP.append(prop)
            elif prop.mattypeid in [48, 99]:
                listGLS.append(prop)
            else:
                listmat.append(prop)
        
        if self.onelist:
            allmat.extend(listDSP)
            allmat.extend(listDVP)
            allmat.extend(listMDF)
            allmat.extend(listGLS)
            allmat.extend(listmat)
            allmat.sort(key=lambda x: [x.mattypeid,x.thickness], reverse=True)
            self.rep_pan('Деталировка', allmat, tpp)
            return
        
        if listDSP:
            listDSP.sort(key=lambda x: x.thickness, reverse=True)
            self.rep_pan('ДСП', listDSP, tpp)
    
        if listMDF:
            listMDF.sort(key=lambda x: x.thickness, reverse=True)
            self.rep_pan('МДФ', listMDF, tpp)
    
        if listDVP:
            listDVP.sort(key=lambda x: x.thickness, reverse=True)
            self.rep_pan('ДВП', listDVP, tpp)
    
        if listGLS:
            listGLS.sort(key=lambda x: x.thickness, reverse=True)
            self.rep_pan('Стекло', listGLS, tpp)
    
        if listmat:
            listmat.sort(key=lambda x: [x.mattypeid,x.thickness], reverse=True)
            self.rep_pan('Прочее', listmat, tpp)

def start(fileDB, projreppath, project, ph, onelist):

    # Подключаемся к базе выгрузки
    db = k3Report.db.DB()
    db.open(fileDB)

    rep = Report(db, ph, onelist)
    rep.makereport()
    db.close()
    file = os.path.join(projreppath, '{}.xlsx'.format(project))
    rep.xl.save(file)
    os.startfile(file)
    return True


if __name__=='__main__':

    fileDB = r'd:\К3\Самара\Самара черновик\249\249.mdb'
    projreppath = r'd:\К3\Самара\Самара черновик\249\Reports'
    ph = '8-999-8791288'
    onelist = 1
    project = "Деталировка заготовок"
    start(fileDB, projreppath, project, ph, onelist)

